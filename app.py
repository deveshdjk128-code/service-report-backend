from flask import Flask, request, send_file
from openpyxl import load_workbook
import tempfile
import os

app = Flask(__name__)

@app.route("/")
def index():
    return "Service Report Backend Running ✅"

@app.route("/generate", methods=["POST"])
def generate_excel():
    data = request.json

    # Load your Excel template (will add file next step)
    wb = load_workbook("template.xlsx")
    ws_list = wb.worksheets  # ALL sheets

    def replace_token(token, value):
        for ws in ws_list:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == token:
                        cell.value = value or ""

    # ---------------- BASIC TOKENS ----------------
    replace_token("{{ DATE}}", data.get("date"))
    replace_token("{{ CUSTOMER NAME}}", data.get("customerName"))
    replace_token("{{ Customer ADDRESS}}", data.get("customerAddress"))
    replace_token("{{ CUSTOMER CONTACT PERSON}}", data.get("contactPerson"))
    replace_token("{{ SERVICE TYPE}}", data.get("serviceType"))

    # ---------------- MACHINE ----------------
    replace_token("{{ MACHINE TYPE}}", data.get("machineType"))
    replace_token("{{ MACHINE S/NO}}", data.get("machineSno"))
    replace_token("{{ MACHINE CODE}}", data.get("machineCode"))
    replace_token("{{ INSTALLATION DATE}}", data.get("installationDate"))

    # ---------------- REPORT ----------------
    replace_token("{{ Report No.}}", data.get("reportNo"))
    replace_token("{{ REPORT NAME}}", data.get("reportName"))
    replace_token("{{ TYPE}}", data.get("type"))

    # ---------------- ACTIVITIES (1–8) ----------------
    for i in range(1, 9):
        replace_token(f"{{{{ ACTIVITY{i}}}}}", data.get(f"activity{i}", ""))
        replace_token(f"{{{{ STATUS{i}}}}}", data.get(f"status{i}", ""))
        replace_token(f"{{{{ JOB BY{i}}}}}", data.get(f"jobBy{i}", ""))

    # ---------------- PARTS ----------------
    for i in range(1, 5):
        replace_token(f"{{{{PART {i}}}}}", data.get(f"part{i}", ""))
        replace_token(f"{{{{PART{i} QTY}}}}", data.get(f"qty{i}", ""))

    # ---------------- TIME ----------------
    replace_token("{{ START TIME}}", data.get("startTime"))
    replace_token("{{ FINISHED TIME}}", data.get("finishTime"))

    # ---------------- JOB & SIGN ----------------
    replace_token("{{ JOB DONE BY}}", data.get("jobDoneBy"))
    replace_token("{{ VERIFIED BY CCP}}", data.get("verifiedByCCP"))
    replace_token("{{ SIGNATURE}}", data.get("signature"))

    # Save temporary Excel
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="service_report.xlsx"
    )

if __name__ == "__main__":
    app.run(debug=True)
